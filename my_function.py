import openpyxl
import pickle

def create_excel():
    return openpyxl.Workbook()

def save_content(data):
    woorkbook = create_excel()
    woorkbook.create_sheet('crawl_data') 
    sheet1 = woorkbook['crawl_data']
    for i in range(len(data)):
        sheet1.cell(row=i+1, column=1).value = i+1
        sheet1.cell(row=i+1, column=2).value = data[i]
    
    save_excel('crawl_data',woorkbook)

def save_excel(name,wb):
    wb.save('excel_logs/'+name+'.xlsx')

def save_list(name,l):
    with open("data_logs/"+name+".txt", "wb") as fp:   #Pickling
        pickle.dump(l, fp)

def load_list(name):
    with open("data_logs/"+name+".txt", "rb") as fp:   # Unpickling
        b = pickle.load(fp)
    print("Success to load.....")
    return b